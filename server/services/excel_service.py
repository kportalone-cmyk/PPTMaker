"""
XLSX 생성 서비스 - openpyxl 기반 스프레드시트 생성

generated_excel 컬렉션의 데이터를 .xlsx 파일로 변환합니다.
차트는 openpyxl 네이티브 차트로 생성하여 OnlyOffice에서 편집 가능합니다.
"""

import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart,
    DoughnutChart, RadarChart, Reference, Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from services.mongo_service import get_db
from bson import ObjectId
from config import settings


async def generate_xlsx(project_id: str) -> str:
    """generated_excel 데이터를 .xlsx 파일로 변환

    Returns:
        생성된 파일의 상대 경로 (예: /uploads/generated/xxx.xlsx)
    """
    db = get_db()
    excel_data = await db.generated_excel.find_one({"project_id": project_id})
    if not excel_data:
        raise ValueError("생성된 엑셀 데이터가 없습니다")

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    sheets = excel_data.get("sheets", [])
    if not sheets:
        raise ValueError("시트 데이터가 없습니다")

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for sheet_data in sheets:
        sheet_name = sheet_data.get("name", "Sheet")[:31]  # Excel 시트명 최대 31자
        ws = wb.create_sheet(title=sheet_name)

        columns = sheet_data.get("columns", [])
        rows = sheet_data.get("rows", [])

        if not columns and not rows:
            continue

        # 헤더 행
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 행
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                if col_idx > len(columns):
                    break
                # 숫자 타입 변환 시도
                value = _try_convert_value(cell_value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                # 짝수 행 배경색
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # 자동 열 너비 계산
        for col_idx in range(1, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    # 한글 문자는 약 2배 너비
                    val_str = str(cell_val)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_len = max(max_len, char_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)

        # 헤더 행 고정
        ws.freeze_panes = "A2"

        # 네이티브 차트 추가 (OnlyOffice 편집 가능)
        _add_native_charts_to_sheet(ws, sheet_data)

    # 파일 저장
    output_dir = os.path.join(settings.UPLOAD_DIR, "generated")
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{uuid.uuid4().hex}.xlsx"
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)

    return f"/uploads/generated/{filename}"


def _try_convert_value(value):
    """문자열을 적절한 타입으로 변환 시도"""
    if value is None or value == "":
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        # 정수 시도
        try:
            return int(value)
        except (ValueError, TypeError):
            pass
        # 실수 시도
        try:
            return float(value)
        except (ValueError, TypeError):
            pass
        # 퍼센트 (예: "85%")
        if value.endswith("%"):
            try:
                return float(value[:-1]) / 100
            except (ValueError, TypeError):
                pass
    return value


# ============ 차트 자동 생성 ============

def auto_generate_chart_definition(sheet_data: dict, chart_type: str = "bar", title: str = None) -> dict | None:
    """시트 데이터에서 자동으로 차트 정의 생성 (LLM 없이)

    Args:
        sheet_data: {"name", "columns", "rows"} 형식의 시트 데이터
        chart_type: bar, line, pie, area, scatter, doughnut, radar
        title: 차트 제목 (None이면 자동 생성)

    Returns:
        차트 정의 dict 또는 None (데이터 부족 시)
    """
    columns = sheet_data.get("columns", [])
    rows = sheet_data.get("rows", [])

    if not columns or not rows:
        return None

    # 라벨 컬럼: 첫 번째 컬럼
    labels_column = 0

    # 숫자 데이터가 있는 컬럼을 시리즈로 선택
    series = []
    for i in range(len(columns)):
        if i == labels_column:
            continue
        numeric_count = 0
        for row in rows:
            if i < len(row):
                val = row[i]
                if isinstance(val, (int, float)):
                    numeric_count += 1
                elif isinstance(val, str):
                    try:
                        float(val.replace(",", ""))
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass
        if numeric_count > 0:
            series.append({"name": columns[i], "column": i})

    if not series:
        return None

    # pie/doughnut은 시리즈 1개만
    if chart_type in ("pie", "doughnut"):
        series = series[:1]

    sheet_name = sheet_data.get("name", "Sheet")
    chart_title = title or sheet_name

    return {
        "type": chart_type,
        "title": chart_title,
        "data_range": {
            "labels_column": labels_column,
            "series": series,
            "row_start": 0,
            "row_end": len(rows) - 1,
        },
        "options": {
            "stacked": False,
            "show_legend": len(series) > 1,
        },
    }


# ============ openpyxl 네이티브 차트 (OnlyOffice 편집 가능) ============

# 색상 팔레트
_CHART_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000",
    "5B9BD5", "70AD47", "264478", "9B59B6",
]


def _add_native_charts_to_sheet(ws, sheet_data: dict):
    """openpyxl 네이티브 차트를 시트에 추가 (OnlyOffice에서 편집 가능)"""
    charts = sheet_data.get("charts", [])
    if not charts:
        return

    rows = sheet_data.get("rows", [])
    columns = sheet_data.get("columns", [])
    num_data_rows = len(rows)
    num_cols = len(columns)

    if num_data_rows == 0 or num_cols == 0:
        return

    for chart_idx, chart_def in enumerate(charts[:3]):
        chart_type = chart_def.get("type", "bar")
        if chart_type not in ("bar", "line", "pie", "area", "scatter", "doughnut", "radar"):
            continue

        try:
            chart_obj = _create_native_chart(ws, chart_def, num_data_rows, num_cols, columns)
            if chart_obj:
                # 데이터 아래 3행, 차트 간 17행 간격
                anchor_row = num_data_rows + 3 + (chart_idx * 17)
                ws.add_chart(chart_obj, f"A{anchor_row}")
        except Exception as e:
            print(f"[Excel] 네이티브 차트 생성 실패 (chart_idx={chart_idx}): {e}")
            continue


def _create_native_chart(ws, chart_def: dict, num_data_rows: int, num_cols: int, columns: list):
    """차트 정의로부터 openpyxl 차트 객체 생성 (OnlyOffice 호환)

    핵심 호환성 포인트:
    1. titles_from_data=True + 헤더 행 포함 → 셀 참조 시리즈 라벨 (strRef)
    2. x_axis.axPos="b", y_axis.axPos="l" → 축 위치 명시적 설정
    3. varyColors=False → 시리즈별 색상 (슬라이스별 아님)
    """
    chart_type = chart_def.get("type", "bar")
    title = chart_def.get("title", "")
    data_range = chart_def.get("data_range", {})
    series_defs = data_range.get("series", [])
    labels_col = data_range.get("labels_column", 0)
    row_start = data_range.get("row_start", 0)
    row_end = data_range.get("row_end")
    options = chart_def.get("options", {})

    if row_end is None:
        row_end = num_data_rows - 1
    row_end = min(row_end, num_data_rows - 1)

    if row_start > row_end or not series_defs:
        return None

    # 헤더 포함 데이터 범위 (titles_from_data=True 용)
    # 헤더: row 1, 데이터: row 2 ~ row_end+2
    header_row = 1
    data_max_row = row_end + 2
    num_categories = data_max_row - header_row  # 헤더 제외한 데이터 행 수

    # 카테고리 라벨 (헤더 제외, 데이터만)
    labels_col_idx = labels_col + 1  # 1-based
    cats = Reference(ws, min_col=labels_col_idx, min_row=header_row + 1, max_row=data_max_row)

    # 데이터 컬럼 인덱스 (1-based)
    data_cols = []
    for s_def in series_defs:
        col_idx = s_def.get("column", 1) + 1
        if 1 <= col_idx <= num_cols:
            data_cols.append(col_idx)
    if not data_cols:
        return None

    # 차트 타입별 객체 생성
    if chart_type == "bar":
        chart_obj = BarChart()
        chart_obj.type = "col"
        chart_obj.grouping = "stacked" if options.get("stacked") else "clustered"
        chart_obj.varyColors = False
        if options.get("stacked"):
            chart_obj.overlap = 100
    elif chart_type == "line":
        chart_obj = LineChart()
        chart_obj.varyColors = False
    elif chart_type == "pie":
        chart_obj = PieChart()
    elif chart_type == "doughnut":
        chart_obj = DoughnutChart()
    elif chart_type == "area":
        chart_obj = AreaChart()
        chart_obj.grouping = "stacked" if options.get("stacked") else "standard"
        chart_obj.varyColors = False
    elif chart_type == "scatter":
        chart_obj = ScatterChart()
        chart_obj.varyColors = False
    elif chart_type == "radar":
        chart_obj = RadarChart()
    else:
        return None

    chart_obj.title = title
    chart_obj.width = max(18, min(num_categories * 2, 30))
    chart_obj.height = 12

    # 시리즈 추가 - titles_from_data=True 패턴 (셀 참조 라벨 생성)
    if chart_type == "scatter":
        # scatter는 xVal/yVal 구조
        for s_idx, col_idx in enumerate(data_cols):
            x_ref = Reference(ws, min_col=labels_col_idx, min_row=header_row + 1, max_row=data_max_row)
            y_ref = Reference(ws, min_col=col_idx, min_row=header_row + 1, max_row=data_max_row)
            series_name = columns[col_idx - 1] if col_idx - 1 < len(columns) else f"Series {s_idx}"
            s = Series(y_ref, xvalues=x_ref, title=series_name)
            chart_obj.series.append(s)
            color = _CHART_COLORS[s_idx % len(_CHART_COLORS)]
            chart_obj.series[-1].graphicalProperties.solidFill = color
    else:
        # 연속 컬럼인 경우 한 번에 추가 (가장 깨끗한 XML 생성)
        data_cols_sorted = sorted(data_cols)
        is_contiguous = (data_cols_sorted == list(range(data_cols_sorted[0], data_cols_sorted[-1] + 1)))

        if is_contiguous:
            data_ref = Reference(ws,
                                 min_col=data_cols_sorted[0],
                                 max_col=data_cols_sorted[-1],
                                 min_row=header_row,
                                 max_row=data_max_row)
            chart_obj.add_data(data_ref, titles_from_data=True)
        else:
            # 비연속 컬럼: 개별 추가 (헤더 포함)
            for col_idx in data_cols:
                data_ref = Reference(ws, min_col=col_idx, min_row=header_row, max_row=data_max_row)
                chart_obj.add_data(data_ref, titles_from_data=True)

        # 카테고리 설정
        if chart_type not in ("pie", "doughnut"):
            chart_obj.set_categories(cats)
        else:
            chart_obj.set_categories(cats)

        # 색상 적용
        for s_idx, s in enumerate(chart_obj.series):
            color = _CHART_COLORS[s_idx % len(_CHART_COLORS)]
            if chart_type not in ("pie", "doughnut"):
                s.graphicalProperties.solidFill = color

    # 축 위치 명시적 설정 (OnlyOffice 호환 핵심)
    if chart_type in ("bar", "line", "area"):
        # 세로 막대: 카테고리축=하단(b), 값축=좌측(l)
        chart_obj.x_axis.axPos = "b"
        chart_obj.y_axis.axPos = "l"
        chart_obj.x_axis.tickLblPos = "nextTo"
        chart_obj.y_axis.tickLblPos = "nextTo"
        chart_obj.x_axis.delete = False
        chart_obj.y_axis.delete = False
        chart_obj.y_axis.numFmt = "#,##0"
        chart_obj.y_axis.majorGridlines = None  # 깔끔한 표시

    # radar 축 설정
    if chart_type == "radar":
        chart_obj.set_categories(cats)

    # pie/doughnut 라벨
    if chart_type in ("pie", "doughnut") and chart_obj.series:
        dl = DataLabelList()
        dl.showPercent = True
        dl.showCatName = True
        dl.showVal = False
        chart_obj.dataLabels = dl

    # 범례
    if not options.get("show_legend", True) and chart_type not in ("pie", "doughnut"):
        chart_obj.legend = None

    return chart_obj
